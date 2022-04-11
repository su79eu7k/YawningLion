import os
import openpyxl as ox
import xlwings as xw


def xw_load_workbooks(filepath):
    return xw.Book(filepath)


def xw_activate_workbook(func):
    def wrapper(workbook_to_activate):
        workbook_to_activate.activate()
        return func(workbook_to_activate)
    return wrapper


@xw_activate_workbook
def xw_get_selection(workbook):
    return workbook.selection.get_address(False, False, False, False)


if __name__ == '__main__':
    test_file = 'D:/Localhome/sekim/OneDrive - ZF Friedrichshafen AG/Desktop/NPV concept.xlsx'
    # ox_load_workbook(test_file)

    wb = xw_load_workbooks(test_file)
    print(xw_get_selection(wb))

